import os
import xml.etree.ElementTree as ET
import unicodedata
from tkinter import Tk, filedialog
import xlsxwriter
import sys
import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QFileDialog, QProgressBar, QLabel
from collections import deque

#OBJEKTUMOK########################################################################################################################
###################################################################################################################################
class Device:
    def __init__(self, id, description, típus, elnevezés, SN, saját_cím, fogadó_eszköz, kommunikáció, fogadó_interface, hálózatazonosító_frekvencia):
        self.id = id
        self.description = description
        self.típus = típus
        self.elnevezés = elnevezés
        self.SN = SN
        self.saját_cím = saját_cím
        self.fogadó_eszköz = fogadó_eszköz
        self.kommunikáció = kommunikáció
        self.fogadó_interface = fogadó_interface
        self.hálózatazonosító_frekvencia = hálózatazonosító_frekvencia
        self.connected_devices = {}

class Group:
    def __init__(self, id, típus, elnevezés, closed = False):
        self.id = id
        self.típus = típus
        self.elnevezés = elnevezés
        self.closed = closed
        self.content = {}


#FÜGGVÉNYEK###########################################################################################################################
######################################################################################################################################
def ignorator_comparator(str1, str2):
    str1_normalized = unicodedata.normalize('NFKD', str1).encode('ASCII', 'ignore').decode('utf-8')
    str2_normalized = unicodedata.normalize('NFKD', str2).encode('ASCII', 'ignore').decode('utf-8')

    str1_normalized = str1_normalized.replace(" ", "")
    str2_normalized = str2_normalized.replace(" ", "")

    str1_normalized = str1_normalized.replace("_", "")
    str2_normalized = str2_normalized.replace("_", "")

    str1_normalized = str1_normalized.replace("-", "")
    str2_normalized = str2_normalized.replace("-", "")

    return str1_normalized.lower() == str2_normalized.lower()

#PARSING########################################################################################################################
################################################################################################################################
def get_node_keys(keys):
    key_ids = []
    for key in keys:
        if key.get('attr.name') and not ignorator_comparator(key.get('attr.name'), 'url') and not None and key.get("for") == "node":
            items = [key.get('id'),key.get('attr.name')]
            key_ids.append(items)
    return key_ids

def get_edge_keys(keys):
    key_ids = []
    for key in keys:
        if key.get('attr.name') and not ignorator_comparator(key.get('attr.name'), 'url') and not None and key.get("for") == "edge":
            items = [key.get('id'),key.get('attr.name')]
            key_ids.append(items)
    return key_ids


def get_element_depth(root, element_id, current_depth=0):
    if root.get('id') == element_id:
        return current_depth
    for child in root:
        depth = get_element_depth(child, element_id, current_depth + 1)
        if depth is not None:
            return depth
    return None


def get_specific_data(node, data_name, keys):
    for key in keys:
        if ignorator_comparator(key[1], data_name):
            my_id = key[0]        
    for data in node.iter("{http://graphml.graphdrawing.org/xmlns}data"):
        if data.get('key') == my_id:
            return data.text
        elif data.get('key') == "d10" or data.get('key') == "d17":
            return None


def reverse_search(item, filename, root_group):
    if isinstance(item, Device):
        root = ET.parse(filename).getroot()
        for node in root.iter("{http://graphml.graphdrawing.org/xmlns}node"):
            if item.id == node.get('id'):
                return node
    else:
        id = item.get('id')
        queue = deque([root_group])
        while queue:
            group = queue.popleft()
            if id in group.content and isinstance(group.content[id], Device):
                return group.content[id]
            for item in group.content.values():
                if isinstance(item, Group):
                    queue.append(item)
    return None
        


#STRUKTÚRA KIALAKÍTÁS################################################################################################################
######################################################################################################################################
def comms(edge, keys):
    if edge.find(".//y:GenericEdge", namespaces={"y": "http://www.yworks.com/xml/graphml"}) is not None:
                    if edge.find(".//y:LineStyle", namespaces={"y": "http://www.yworks.com/xml/graphml"}).get("type") == "dashed":
                                return "wifi" 
                    if edge.find(".//y:LineStyle", namespaces={"y": "http://www.yworks.com/xml/graphml"}).get("type") == "dashed_dotted":
                                return "gsm"
                    if edge.find(".//y:LineStyle", namespaces={"y": "http://www.yworks.com/xml/graphml"}).get("type") == "line":
                                return "ethernet"
    if edge.find(".//y:LineStyle", namespaces={"y": "http://www.yworks.com/xml/graphml"}).get("type") == "dashed":
        return "rádiós soros port" 
    if edge.find(".//y:LineStyle", namespaces={"y": "http://www.yworks.com/xml/graphml"}).get("type") == "dashed_dotted":
        return "rádiós impulzus"
    if edge.find(".//y:LineStyle", namespaces={"y": "http://www.yworks.com/xml/graphml"}).get("type") == "dotted":
        return "LoRa"
    if get_specific_data(edge, "Kommunikáció Módja", get_edge_keys(keys)) is not None:
        return get_specific_data(edge, "Kommunikáció Módja", get_edge_keys(keys))        


def get_all_devices(filename):
    tree = ET.parse(filename)
    root = tree.getroot()
    keys = root.findall(".//{http://graphml.graphdrawing.org/xmlns}key")
    root_group = Group('root', 'root', filename[:-8])
    groups = {'root': root_group}
    for node in root.iter('{http://graphml.graphdrawing.org/xmlns}node'):
        if 'yfiles.foldertype' in node.attrib and (node.attrib['yfiles.foldertype'] == 'group' or node.attrib['yfiles.foldertype'] == 'folder'):
            group = Group(id=node.get("id"),
                        típus=get_specific_data(node, "Típus", get_node_keys(keys)),
                        elnevezés=get_specific_data(node, "Elnevezés", get_node_keys(keys)),)
            if node.attrib['yfiles.foldertype'] == 'folder':
                group.closed = True
            groups[node.get('id')] = group
            if len(node.get('id')) > 5:
                parent_id = node.get('id').rsplit('::', 1)[0]
                if parent_id in groups:
                    groups[parent_id].content[group.id] = group
            else:
                parent_id = "root"
                if parent_id in groups:
                    groups[parent_id].content[group.id] = group
        elif get_specific_data(node, "description", get_node_keys(keys)) is not None:
            device = Device(id=node.get("id"), 
                            description=get_specific_data(node, "description", get_node_keys(keys)),
                            típus=get_specific_data(node, "Típus", get_node_keys(keys)),
                            elnevezés=get_specific_data(node, "Elnevezés", get_node_keys(keys)),
                            SN=get_specific_data(node, "SN", get_node_keys(keys)),
                            saját_cím=get_specific_data(node, "Saját Cím", get_node_keys(keys)),
                            fogadó_eszköz=None,
                            kommunikáció=None,
                            fogadó_interface=None,
                            hálózatazonosító_frekvencia=None,)
            if len(node.get('id')) > 2:
                parent_id = node.get('id').rsplit('::', 1)[0]
                if parent_id in groups:
                    groups[parent_id].content[device.id] = device
            else:
                parent_id = "root"
                if parent_id in groups:
                    groups[parent_id].content[device.id] = device
    connections(filename, root_group)
    return root_group

def count_devices_by_type(group, filename):
    counts = {}
    for id, item in group.content.items():
        if isinstance(item, Device):
            if item.típus == None:
                item.típus = item.description
            if item.típus not in counts:
                counts[item.típus] = 0
            if item.description == "Áramváltó":
                node = reverse_search(item, filename, group)
                fill_color = node.find(".//y:Fill", namespaces={"y": "http://www.yworks.com/xml/graphml"}).get("color")
                if fill_color is not None and len(fill_color) > 7 and fill_color[:8] == '#0000000':
                        counts[item.típus] += int(fill_color[-1])
                else:
                    counts[item.típus] += 1
            else:    
                counts[item.típus] += 1
        elif isinstance(item, Group):
            sub_counts = count_devices_by_type(item, filename)
            for típus, count in sub_counts.items():
                if típus not in counts:
                    counts[típus] = 0
                counts[típus] += count
    return counts

def connections(filename, root_group, parent=None, visited=None, visited_edges=None):
    root = ET.parse(filename).getroot()
    edges = root.findall(".//{http://graphml.graphdrawing.org/xmlns}edge")
    nodes = root.findall(".//{http://graphml.graphdrawing.org/xmlns}node")
    keys = root.findall(".//{http://graphml.graphdrawing.org/xmlns}key")
    if visited is None:
        visited = set()
        visited_edges = set()
        for node in nodes:
            if get_specific_data(node, "description", get_node_keys(keys)) == "Szerver":
                parent = reverse_search(node, filename, root_group)
    if parent.id in visited:
        return
    visited.add(parent.id)
    for edge in edges:
        if edge.get("source") == parent.id and edge.get('id') not in visited_edges:
             visited_edges.add(edge.get('id'))
             for node in nodes:
                  if node.get("id") == edge.get("target"):
                        device = reverse_search(node, filename, root_group)
                        if device is None:
                            return
                        device.fogadó_eszköz=get_specific_data(node, "Típus", get_node_keys(keys))
                        device.kommunikáció=comms(edge, keys)
                        device.fogadó_interface=get_specific_data(edge, "Fogadó interface", get_edge_keys(keys))
                        device.hálózatazonosító_frekvencia=get_specific_data(edge, "Hálózatazonosító/Frekvencia", get_edge_keys(keys))
                        if device.id not in visited:
                            connections(filename, root_group, device, visited, visited_edges)
                            parent.connected_devices[device.id] = device
                        else:
                            return
        if edge.get("target") == parent.id and edge.get('id') not in visited_edges:
             visited_edges.add(edge.get('id'))
             for node in nodes:
                  if node.get("id") == edge.get("source"):
                        device = reverse_search(node, filename, root_group)
                        if device is None:
                            return
                        device.fogadó_eszköz=get_specific_data(reverse_search(parent, filename, root_group), "Típus", get_node_keys(keys))
                        device.kommunikáció=comms(edge, keys)
                        device.fogadó_interface=get_specific_data(edge, "Fogadó interface", get_edge_keys(keys))
                        device.hálózatazonosító_frekvencia=get_specific_data(edge, "Hálózatazonosító/Frekvencia", get_edge_keys(keys))
                        if parent.típus and "remx" in parent.típus.lower() and device.description == "Áramváltó":
                            device.típus += "/50A"
                        if device.id not in visited:
                            connections(filename, root_group, device, visited, visited_edges)
                            parent.connected_devices[device.id] = device
                        else:
                            return


def find_devices(server, sought_description):
    stack = []
    found_devices = []
    stack.append(server)
    while stack:
        device = stack.pop()
        if device.description == sought_description:
            found_devices.append(device)
        for connected_device in device.connected_devices.values():
            stack.append(connected_device)
    return found_devices


#XLSX KÉSZÍTÉS#####################################################################################################################################
######################################################################################################################################
def create_xls(filename):
    root = Tk()
    root.withdraw()
    default_filename = filename[:-8].replace('_PROXY', '') + '.xlsx'
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_filename, filetypes=[("Excel Workbook", "*.xlsx")])
    if not file_path:
        return
    
    root_group = get_all_devices(filename)
    counts = count_devices_by_type(root_group, filename)
    workbook = xlsxwriter.Workbook(file_path)
    worksheet1 = workbook.add_worksheet('Összeszerelés')
    worksheet2 = workbook.add_worksheet('Anyaglista')

    attr_name_format = workbook.add_format({'border': 2, 'bold': True, 'bg_color': '#7AEC92'})
    group_format = workbook.add_format({'border': 2, 'bold': True, 'bg_color': '#EDED79'})
    red_bg_format = workbook.add_format({'bg_color': 'red'})
    attr_value_format = workbook.add_format({'border': 1, 'bg_color': '#BDF5C9'})


    row = 0

    def traverse(group, level=0):
        nonlocal row
        if group.closed:
            return
        if level > 0:
            group_info = f"{group.típus or ''} || {group.elnevezés or ''}"
            worksheet1.merge_range(row, 0, row, 10 - min(level, 3), group_info, group_format)
            row += 1
            if any(isinstance(item, Device) for item in group.content.values()):
                headers = ['Típus', 'Elnevezés', 'SN', 'Saját Cím', 'Fogadó Eszköz', 'Kommunikáció', 'Fogadó Interface', 'Hálózatazonosító Frekvencia']
                worksheet1.write_row(row, 0, headers, attr_name_format)
                row += 1
        for item in group.content.values():
            if isinstance(item, Device):
                worksheet1.write_row(row, 0, [item.típus, item.elnevezés, item.SN, item.saját_cím, item.fogadó_eszköz, item.kommunikáció, item.fogadó_interface, item.hálózatazonosító_frekvencia], attr_value_format)
                row += 1
        for item in group.content.values():
            if isinstance(item, Group):
                traverse(item, level+1)

    traverse(root_group)

    row = 0
    col = 0

    worksheet2.write(row, col, 'Eszköz', group_format)
    worksheet2.write(row, col + 1, 'Darabszám', group_format)

    row += 1

    for key, value in counts.items():
        worksheet2.write(row, col, key, attr_value_format)
        worksheet2.write(row, col + 1, value, attr_value_format)
        row += 1


    workbook.close()
    return file_path


#GUI##################################################################################################################################
######################################################################################################################################
class Sorter(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        
    def initUI(self):
        self.setGeometry(400, 400, 500, 300)
        self.setWindowTitle('XLSX Maker')
        
        self.choose_file_btn = QPushButton('Choose File', self)
        self.choose_file_btn.move(10, 10)
        self.choose_file_btn.clicked.connect(self.chooseFile)

        self.choose_file_btn_label = QLabel('                                                                                                                ', self)
        self.choose_file_btn_label.move(110, 14)
        
        self.upload_btn = QPushButton('Create XLSX', self)
        self.upload_btn.move(10, 50)
        self.upload_btn.clicked.connect(self.makeXls)

        self.open_file_btn = QPushButton('Open XLSX', self)
        self.open_file_btn.move(10, 150)
        self.open_file_btn.clicked.connect(self.openFile)
        self.open_file_btn.hide()
        
        self.pb = QProgressBar(self)
        self.pb.setGeometry(10, 120, 380, 20)
        
        self.label = QLabel('                                                    ', self)
        self.label.move(10, 90)
        
    def makeXls(self):
        if not hasattr(self, 'file_path'):
            return

        xlsx_file_path = create_xls(self.file_path)
        
        if not xlsx_file_path:
            return

        xlsx_filename = os.path.basename(xlsx_file_path)
        self.open_file_btn.setText(f'Open {xlsx_filename}')
        self.open_file_btn.show()

        self.pb.setValue(100)
        self.label.setText('XLSX created successfully!')    

    def openFile(self):
        if not hasattr(self, 'file_path'):
            return

        filename = os.path.splitext(self.file_path)[0] + '.xlsx'

        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook['Összeszerelés']
        mistakes = False
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == '?':
                    mistakes = True
                    break
            if mistakes:
                break

        if mistakes:
            import tkinter as tk
            from tkinter import messagebox

            root = tk.Tk()
            root.withdraw()

            answer = messagebox.askyesno("Hiányok", "Hiányosság van a vázlaton\nÍgy is meg szeretné nyitni a fájlt?")
            if answer:
                os.startfile(filename)
            else:
                sys.exit()
        else:
            os.startfile(filename)
    
    def chooseFile(self):
        self.file_path, _ = QFileDialog.getOpenFileName(self, 'Choose File', '', 'XML Files (*.graphml)')
        
        #self.proxy_file_path = create_proxy(self.file_path)
        filename = os.path.basename(self.file_path)
        
        self.choose_file_btn_label.setText(filename)
    
    

if __name__ == '__main__':
    app = QApplication(sys.argv)
    sorter = Sorter()
    sorter.show()
    sys.exit(app.exec_())
